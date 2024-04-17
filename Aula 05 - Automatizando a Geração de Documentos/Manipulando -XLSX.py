# Instando a biblioteca openpyxl
# pip install openpyxl

'''
from openpyxl import load_workbook
arquivo = 'planilha.xlsx'
workbook=load_workbook(arquivo)
planilha=workbook.active

for linha in planilha.iter_rows():
    for celula in linha:
        print(str(celula.value).ljust(20), end='') #\t - espaço
    print()

'''

#exibir uma informção usando referência de célula
import openpyxl

planilha=openpyxl.load_workbook('planilha.xlsx')
aba=planilha['PRODUTOS']
print(aba['a2'].value)
print(aba['b2'].value)
print(aba['a3'].value)
print(aba['b3'].value)




